import openpyxl
import numpy
read_delay_n_1_to_10=[]
write_delay_n_1_to_10=[]

def cal_delay(strings_name):
    wb=openpyxl.load_workbook(strings_name)
    sheet=wb.get_sheet_by_name(wb.get_sheet_names()[0])
    pe_sink_list=[]

    for i in range(2,57,3):
        temp=[]
        for j in range(3000,3602,1):
            if(sheet.cell(j,i).value!=0):
                temp.append(sheet.cell(j,i).value)
        pe_sink_list.append(numpy.mean(temp))
    print("write request delay")
    write_delay_n_1_to_10.append((pe_sink_list[14]+pe_sink_list[16]+pe_sink_list[18])/3)
    print((pe_sink_list[14]+pe_sink_list[16]+pe_sink_list[18])/3)   
    print("read request delay")
    read_request_delay=0
    pe_to_mm=0
    mm_to_pe=0
    for i in range(0,13,1):
        mm_to_pe+=pe_sink_list[i]
    mm_to_pe=mm_to_pe/13
    pe_to_mm=(pe_sink_list[13]+pe_sink_list[15]+pe_sink_list[17])/3
    read_request_delay=mm_to_pe+pe_to_mm
    read_delay_n_1_to_10.append(read_request_delay)
    print(read_request_delay)
    return

for i in range(1,11,1):
    file_name=r'C:\Users\tes\Desktop\opnet_data\without_line_factor\uuid1_proto-scenario1-DES-1__T_'+str(i)+r'.xlsx'
    cal_delay(file_name)
print(write_delay_n_1_to_10)
print(read_delay_n_1_to_10)

